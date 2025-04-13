import pandas as pd
import openpyxl
from string import ascii_uppercase
from openpyxl.styles import Alignment, PatternFill, Font
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.utils import get_column_letter
from openpyxl.styles.borders import Border, Side


# function to generate the plate layout
def generate_plate_layout(
    input_file, sheet, col, row, extraction_replicates, output_folder, project_name
):
    # open the excel file with the sheetname
    wb = openpyxl.load_workbook(input_file, data_only=True)

    # open the worksheet
    ws = wb[sheet]

    # convert the inputs to the correct format
    letter_to_numeric = list(ascii_uppercase) + [
        i + j for i in ascii_uppercase for j in ascii_uppercase
    ]
    letter_to_numeric = dict(
        zip(letter_to_numeric, range(1, len(letter_to_numeric) + 2))
    )

    col = col.capitalize()
    col = letter_to_numeric[col]
    row = int(row)

    # extract the samples
    samples = [ws.cell(row=row, column=col).value for row in range(row, ws.max_row + 1)]
    sample_count = len(samples)
    replicate_numbers = sample_count * list(range(1, extraction_replicates + 1))

    # include the extraction replicates if they are greater than 1
    if extraction_replicates > 1:
        samples = sorted(list(enumerate(samples)) * extraction_replicates)
        samples = [
            "{}_{}".format(sample[1], number)
            for sample, number in zip(samples, replicate_numbers)
        ]

    ## first sample will add with every step in the while loop
    number = 1
    letter = 0
    well = 1
    ncs = [8, 9, 21, 27, 39, 44, 54, 58, 68, 73, 88, 93]
    nc = 1
    output_dict = {}
    letters = list(ascii_uppercase) + [
        i + j for i in ascii_uppercase for j in ascii_uppercase
    ]

    # add a dict with alphanumeric convention
    alpha = [
        "{}{}".format(letter, well)
        for well in range(1, 13)
        for letter in ["A", "B", "C", "D", "E", "F", "G", "H"]
    ]
    alpha = dict(zip(range(1, 97), alpha))

    while samples:
        ## if the plate and the replicate is full jump to plate 3
        if (number - 1) % 96 == 0 and number != 1:
            number += 96
            letter += 2
            well = 1
            nc = 1

        ## if well contains a negative controll, add negative control to the output dict
        if well in ncs:
            ## first replicate
            output_dict[number] = [
                "NC_{}_{}_{}{}".format(
                    nc, project_name, letters[letter], letters[letter + 1]
                ),
                letters[letter],
                alpha[well],
            ]
            output_dict[number + 96] = [
                "NC_{}_{}_{}{}".format(
                    nc, project_name, letters[letter], letters[letter + 1]
                ),
                letters[letter + 1],
                alpha[well],
            ]
            nc += 1
            number += 1
            well += 1
        ## else add a sample
        else:
            sample = samples.pop(0)
            output_dict[number] = ["{}".format(sample), letters[letter], alpha[well]]
            output_dict[number + 96] = [
                "{}".format(sample),
                letters[letter + 1],
                alpha[well],
            ]
            number += 1
            well += 1

    out_df = []

    for key in sorted(output_dict.keys()):
        out_df.append(output_dict[key])

    out_df = pd.DataFrame(
        out_df,
        columns=[
            "Sample name",
            "Lysis,\nExtraction,\nPCR plate",
            "Lysis,\nExtraction,\nPCR Well",
        ],
    )

    # order for tubes
    tube_order = [
        "{}{}".format(letter, number)
        for letter in ["A", "B", "C", "D", "E", "F"]
        for number in reversed(range(1, 9))
    ]

    alpha = [
        "{}{}".format(letter, well)
        for well in range(1, 13)
        for letter in ["A", "B", "C", "D", "E", "F", "G", "H"]
    ]

    # add tube order
    tube_order = dict(zip(alpha, tube_order + tube_order))
    out_df["Lysis\ntube"] = out_df["Lysis,\nExtraction,\nPCR Well"].map(tube_order)

    # add tagging primers
    alpha = dict(zip(alpha, range(1, 97)))
    out_df["aliquoted"] = ""
    out_df["Primer\n2nd PCR"] = out_df["Lysis,\nExtraction,\nPCR Well"].map(alpha)
    out_df["Notes"] = ""
    out_df = out_df.reset_index(names="#")
    out_df["#"] = out_df["#"] + 1

    # create a temporary save before doing some styling
    savename = "{}_plate_layout.xlsx".format(project_name)
    savename = output_folder.joinpath(savename)
    out_df.to_excel(savename, index=False, sheet_name="plate_layout")
    wb.close()

    # use openpyxl to style everything
    wb = openpyxl.load_workbook(savename)
    ws = wb["plate_layout"]

    # add the line breaks for styling
    header = ["A1", "B1", "C1", "D1", "E1", "F1", "G1", "H1"]
    for cell in header:
        ws[cell].alignment = Alignment(wrap_text=True)

    # adjust cell width for all columns
    dim_holder = DimensionHolder(worksheet=ws)
    widths = [5, 25, 12, 12, 8, 12, 10, 25]

    for col, width in zip(range(ws.min_column, ws.max_column + 1), widths):
        if col == 2:
            dim_holder[get_column_letter(col)] = ColumnDimension(
                ws, min=col, max=col, width=width
            )
        else:
            dim_holder[get_column_letter(col)] = ColumnDimension(
                ws, min=col, max=col, width=width
            )

    ws.column_dimensions = dim_holder

    ## add a header
    ws.insert_rows(1)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    ws.cell(row=1, column=1).value = "Plate Layout Project: {}".format(project_name)

    ## add alternating cell styling and borders
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    fill = PatternFill(start_color="00C0C0C0", end_color="00C0C0C0", fill_type="solid")

    for row in range(2, ws.max_row + 1):
        for col in range(1, 9):
            ws.cell(row=row, column=col).border = thin_border
            if row % 2 != 0:
                ws.cell(row=row, column=col).fill = fill

    # styling for the header
    ws.cell(row=1, column=1).font = Font(bold=True)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")

    wb.save(savename)
    wb.close()
