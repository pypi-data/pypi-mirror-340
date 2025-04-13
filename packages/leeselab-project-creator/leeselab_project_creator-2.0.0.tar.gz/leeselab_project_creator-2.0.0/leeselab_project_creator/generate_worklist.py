import openpyxl, sys
import pandas as pd
from openpyxl.styles import Alignment, PatternFill, Font
from openpyxl.styles.borders import Border, Side


# funtion to calculate one worklist per marker
def worklist_per_marker(
    available_primers,
    extraction_plates,
    pcr_replicates,
    primer_name,
    optimal_primers,
    starting_library_number,
    pool_pcr_replicates,
):
    # generate the output dataframe
    output_worklist = pd.DataFrame()
    output_worklist["source plate"] = [
        extraction_plate
        for extraction_plate in extraction_plates
        for _ in range(pcr_replicates)
    ]

    # find the number of pcr plates for this marker
    pcr_plate_count = len(output_worklist)

    # find the number of librarys to distribute primers evenly
    available_primers = available_primers.split(",")
    available_primers = [int(primer) for primer in available_primers]

    # add a switch here if this code works
    if pool_pcr_replicates:
        optimal_primers = [
            primer for primer in optimal_primers for _ in range(pcr_replicates)
        ]
        number_of_primers = len(available_primers) * pcr_replicates
    else:
        number_of_primers = len(available_primers)

    while True:
        if pcr_plate_count <= number_of_primers:
            library_count = 1
            break
        elif pcr_plate_count % number_of_primers == 0:
            library_count = int(pcr_plate_count / number_of_primers)
            break
        else:
            number_of_primers -= 1

    # calculate the number of plates per library
    plates_per_library = int(pcr_plate_count / library_count)

    # generate the library column
    library_column = []

    for _ in range(library_count):
        for _ in range(plates_per_library):
            library_column.append(starting_library_number)
        starting_library_number += 1

    output_worklist["library"] = library_column

    primers = []

    # find the optimal primers for maximum library diversity
    for primer in optimal_primers:
        if len(primers) < plates_per_library:
            if primer in available_primers:  # primer not in primers and
                primers.append(primer)
        else:
            break

    primers = primers * library_count
    primers = [
        "{} - {}".format(primer_name, primer_number) for primer_number in primers
    ]

    output_worklist["tagging primer"] = primers
    output_worklist["1st pcr"] = ""
    output_worklist["clean up"] = ""
    output_worklist["2nd pcr"] = ""
    output_worklist["gel 2nd pcr"] = ""
    output_worklist["normalization"] = ""
    output_worklist["pooling"] = ""

    return (
        output_worklist,
        starting_library_number,
        plates_per_library,
    )


def generate_worklist(
    output_path,
    project,
    available_primers,
    pcr_replicates,
    markers,
    pool_pcr_replicates,
):
    # generate the correct path to the file
    input_file = "{}_plate_layout.xlsx".format(project)
    input_file = output_path.joinpath(input_file)

    # read the data
    data = pd.read_excel(input_file, sheet_name="plate_layout", skiprows=1)

    # collect all worklists here for the final output
    worklists = []

    # collect the plate letters
    plates = list(data["Lysis,\nExtraction,\nPCR plate"].unique())
    general_worklist = pd.DataFrame(
        columns=[
            "plate",
            "aliquoted",
            "lysis",
            "inhibitor removal",
            "lysate distribution",
            "extraction",
            "gel extraction",
        ]
    )

    # general worklist is done
    general_worklist["plate"] = plates

    # optimal primer order for maximizing library diversity
    optimal_primer_order = [
        1,
        5,
        9,
        2,
        6,
        10,
        3,
        7,
        11,
        4,
        8,
        12,
        13,
        17,
        21,
        14,
        18,
        22,
        15,
        19,
        23,
        16,
        20,
        24,
    ]

    # gather the worklists here
    worklists = []

    # calculate everything marker wise and concat the marker dfs in the end
    next_library = 1
    markers = markers.split(",")

    for primer in markers:
        worklist, next_library, plates_per_library = worklist_per_marker(
            available_primers,
            plates,
            pcr_replicates,
            primer,
            optimal_primer_order,
            next_library,
            pool_pcr_replicates,
        )

        worklists.append(worklist)

    # concat the individual worklist to have a working table
    working_table = pd.concat(worklists, axis=0)

    # save both tables to excel to perform styling via openpyxl
    # tables will be saved as sheets to the plate layout
    with pd.ExcelWriter(
        input_file, mode="a", if_sheet_exists="replace", engine="openpyxl"
    ) as writer:
        general_worklist.to_excel(writer, sheet_name="extraction_worklist", index=False)
        working_table.to_excel(writer, sheet_name="pcr_worklist", index=False)

    ## add the styling
    add_styling(input_file, project, plates_per_library)


# function to add styling to the worklists
def add_styling(input_file, project, plates_per_library):

    # open the extraction worklist first
    wb = openpyxl.load_workbook(input_file)
    ws = wb["extraction_worklist"]

    # Iterate over all columns and adjust their widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width

    # add a connected header cell
    ws.insert_rows(1)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
    ws.cell(row=1, column=1).value = "Extraction worklist: {}".format(project)

    # styling for the header
    ws.cell(row=1, column=1).font = Font(bold=True)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")

    # add alternatig colors
    ## add alternating cell styling and borders
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    fill = PatternFill(start_color="00C0C0C0", end_color="00C0C0C0", fill_type="solid")

    for row in range(2, ws.max_row + 1):
        for col in range(1, 8):
            ws.cell(row=row, column=col).border = thin_border
            if row % 2 == 0:
                ws.cell(row=row, column=col).fill = fill

    # merge cells where steps are executed together
    for col in range(2, 6):
        for row in range(3, ws.max_row + 1, 2):
            ws.merge_cells(
                start_row=row, end_row=row + 1, start_column=col, end_column=col
            )

    # add styling for the pcr worklist
    ws = wb["pcr_worklist"]

    # Iterate over all columns and adjust their widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width

    # add a connected header cell
    ws.insert_rows(1)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    ws.cell(row=1, column=1).value = "PCR worklist: {}".format(project)

    # styling for the header
    ws.cell(row=1, column=1).font = Font(bold=True)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")

    # add alternatig colors
    ## add alternating cell styling and borders
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    fill = PatternFill(start_color="00C0C0C0", end_color="00C0C0C0", fill_type="solid")

    for row in range(2, ws.max_row + 1):
        for col in range(1, 10):
            ws.cell(row=row, column=col).border = thin_border
            if row % 2 == 0:
                ws.cell(row=row, column=col).fill = fill

    # merge cells where steps are executed together
    for col in range(8, 10):
        for row in range(3, ws.max_row + 1, plates_per_library):
            ws.merge_cells(
                start_row=row,
                end_row=row + plates_per_library - 1,
                start_column=col,
                end_column=col,
            )

    wb.save(input_file)
    wb.close()
