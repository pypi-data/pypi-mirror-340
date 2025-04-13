from openpyxl.utils import get_column_letter
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.drawing.image import Image
import os
import seaborn as sns

def create_plot(data, x_column, y_column=None, plot_type="scatter", title="", png_output_path="", svg_output_path="", color_map=None):
    """
    Create enhanced plots with support for scatter, box, and histogram plot types.
    """
    plt.figure(figsize=(12, 8))

    # Set color palette
    if color_map is None:
        unique_tasks = data["task"].unique() if "task" in data.columns else []
        palette = sns.color_palette("Set2", n_colors=len(unique_tasks))
        color_map = {task: palette[i % len(palette)] for i, task in enumerate(unique_tasks)}

    try:
        if plot_type == "scatter":
            if x_column not in data or y_column not in data:
                raise ValueError(f"Columns {x_column} or {y_column} not found in data for scatter plot.")
            sns.scatterplot(data=data, x=x_column, y=y_column, hue="task", palette=color_map, alpha=0.7)
            plt.legend(title="Task", fontsize=10, loc="best")
        elif plot_type == "box":
            if x_column not in data or y_column not in data:
                raise ValueError(f"Columns {x_column} or {y_column} not found in data for box plot.")
            sns.boxplot(data=data, x=x_column, y=y_column, hue="task", palette=color_map, linewidth=1.5, dodge=True)
            plt.legend(title="Task", fontsize=10, loc="best")
        elif plot_type == "histogram":
            if x_column not in data.columns:
                raise ValueError(f"Column {x_column} not found in data for histogram.")
            sns.histplot(data[x_column], bins=30, kde=True, color="teal", edgecolor="black", alpha=0.7)
            plt.ylabel("Frequency", fontsize=14)
        else:
            raise ValueError(f"Unsupported plot type: {plot_type}")

        # Titles and Labels
        plt.title(title, fontsize=18, weight="bold")
        plt.xlabel("Training Time (nanoseconds)" if x_column == "duration" else x_column.capitalize(), fontsize=14)
        if y_column:
            plt.ylabel(y_column.capitalize(), fontsize=14)
        plt.xticks(fontsize=12)
        plt.yticks(fontsize=12)

        # Grid and Layout
        plt.grid(True, linestyle="--", alpha=0.5)
        plt.tight_layout()

        # Save the plot in both formats
        plt.savefig(png_output_path, format="png", dpi=300)
        plt.savefig(svg_output_path, format="svg")
        plt.close()
    except Exception as e:
        print(f"Error generating {plot_type} plot: {e}")

def plot_first_epoch_training_time_distribution(data, title, png_output_path, svg_output_path):
    """
    Creates an enhanced distribution plot for the training time of the first epoch.
    """
    first_epoch_data = data[data["epoch"] == 1]

    if first_epoch_data.empty:
        print("No data available for the first epoch training time distribution. Skipping plot generation.")
        return

    plt.figure(figsize=(12, 8))
    sns.histplot(first_epoch_data["duration"], kde=True, bins=20, color="blue", edgecolor="black", alpha=0.7)

    plt.title(title, fontsize=18, weight="bold")
    plt.xlabel("Training Time (nanoseconds)", fontsize=14)
    plt.ylabel("Count", fontsize=14)
    plt.xticks(fontsize=12)
    plt.yticks(fontsize=12)

    plt.grid(True, linestyle="--", alpha=0.7)
    plt.tight_layout()
    plt.savefig(png_output_path, format="png", dpi=300)
    plt.savefig(svg_output_path, format="svg")
    plt.close()

def plot_first_epoch_training_time_distribution_by_model(data, title, png_output_path, svg_output_path):
    """
    Create an insightful plot of model training times for the first epoch, grouped by models (or datasets).
    """
    first_epoch_data = data[data["epoch"] == 1]

    if first_epoch_data.empty:
        print("No data available for the first epoch training time distribution. Skipping plot generation.")
        return

    aggregated_data = first_epoch_data.groupby(["nn", "dataset"])["duration"].median().reset_index()

    plt.figure(figsize=(16, 8))
    sns.barplot(data=aggregated_data, x="nn", y="duration", hue="dataset", palette="Set2", dodge=True)

    plt.yscale("log")
    plt.title(title, fontsize=18, weight="bold")
    plt.xlabel("Model (nn)", fontsize=14)
    plt.ylabel("Training Time (nanoseconds)", fontsize=14)
    plt.xticks(rotation=45, fontsize=12, ha="right")
    plt.legend(title="Dataset", fontsize=10, loc="upper left", bbox_to_anchor=(1, 1))
    plt.grid(linestyle="--", alpha=0.5)

    plt.tight_layout()
    plt.savefig(png_output_path, format="png", dpi=300)
    plt.savefig(svg_output_path, format="svg")
    plt.close()

def export_raw_data_with_plots(filtered_raw_data, output_file, png_dir, svg_dir):
    """
    Export raw data and plots to an Excel file, positioning the plots vertically.
    """
    os.makedirs(png_dir, exist_ok=True)
    os.makedirs(svg_dir, exist_ok=True)

    unique_tasks = filtered_raw_data["task"].unique()
    color_map = {task: color for task, color in zip(unique_tasks, sns.color_palette("tab10", len(unique_tasks)))}

    plot_configs = [
        ("epoch", "accuracy", "scatter", "Accuracy vs Epochs", "accuracy_vs_epochs"),
        ("duration", "accuracy", "scatter", "Accuracy vs Training Time", "accuracy_vs_training_time"),
        ("epoch", "accuracy", "box", "Accuracy Distribution by Epochs", "accuracy_vs_epochs_box"),
        ("accuracy", None, "histogram", "Accuracy Frequency", "accuracy_histogram"),
    ]

    plot_paths = []
    for x, y, plot_type, title, filename in plot_configs:
        png_output_path = os.path.join(png_dir, f"{filename}.png")
        svg_output_path = os.path.join(svg_dir, f"{filename}.svg")
        create_plot(filtered_raw_data, x, y, plot_type, title, png_output_path, svg_output_path, color_map)
        plot_paths.append((title, png_output_path))

    first_epoch_filename = "first_epoch_training_time_distribution_by_model"
    first_epoch_title = "Distribution of Model Training Times (First Epoch)"
    first_epoch_png_path = os.path.join(png_dir, f"{first_epoch_filename}.png")
    first_epoch_svg_path = os.path.join(svg_dir, f"{first_epoch_filename}.svg")
    plot_first_epoch_training_time_distribution_by_model(
        filtered_raw_data, first_epoch_title, first_epoch_png_path, first_epoch_svg_path
    )
    plot_paths.append((first_epoch_title, first_epoch_png_path))

    wb = Workbook()
    ws = wb.active
    ws.title = "Raw Data"

    ws.append(filtered_raw_data.columns.tolist())
    for row in filtered_raw_data.itertuples(index=False):
        ws.append(row)

    start_row = 2
    plot_column_start = len(filtered_raw_data.columns) + 5

    for title, png_path in plot_paths:
        if os.path.exists(png_path):
            ws.cell(row=start_row, column=plot_column_start, value=title)
            img = Image(png_path)
            img.width = 500
            img.height = 350
            img.anchor = f"{get_column_letter(plot_column_start)}{start_row + 1}"
            ws.add_image(img)
            start_row += 20

    wb.save(output_file)
    print(f"Raw data and plots saved to {output_file}")
