from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from openpyxl.drawing.image import Image
import os


def export_to_excel(aggregated_data, output_file, plot_dir, image_width=400, image_height=300):
    """
    Export aggregated data and categorized plots to an Excel file.
    Args:
        aggregated_data (DataFrame): Aggregated statistics.
        output_file (str): Path to save the Excel file.
        plot_dir (str): Directory containing plots.
        image_width (int): Width of embedded images (default: 400).
        image_height (int): Height of embedded images (default: 300).
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Statistics"

    # Add headers
    ws.append(aggregated_data.columns.tolist())

    # Add data rows
    for row in aggregated_data.itertuples(index=False):
        ws.append(row)

    # Define plot categories and their corresponding file patterns
    parallel_categories = {
        "Mean and Std Dev": "{task}_{dataset}_{metric}_mean_std.png",
        "Box Plot": "{task}_{dataset}_{metric}_box.png",
        "Rolling Mean": "{task}_{dataset}_{metric}_rolling_mean.png",
        
    }
    sequential_categories = {
        "Correlation Heatmap": "correlation_heatmap.png",
        "Accuracy vs Training Time": "accuracy_vs_training_time.png",
        "IoU vs Training Time": "iou_vs_training_time.png",
        "Distribution of Model Durations Image classification": "first_epoch_duration_distribution_img-classification.png",
        "Distribution of Model Durations Image segmentation": "first_epoch_duration_distribution_img-segmentation.png",

    }

    # Add parallel layout plots (adjust positions for each category)
    start_row = 2  # Start embedding images below the data headers
    rightmost_column = len(aggregated_data.columns) + 3  # Start 3 columns after the last data column
    for category, file_suffix in parallel_categories.items():
        if category == "Mean and Std Dev":
            category_row = start_row - 1
        elif category == "Box Plot":
            start_row += 3
            category_row = start_row - 1
        elif category == "Rolling Mean":
            start_row += 3
            category_row = start_row - 1

        # Add the category name above the row of images
        ws.cell(row=category_row, column=rightmost_column, value=category)
        col_offset = rightmost_column  # Start placing images

        for metric in aggregated_data['metric'].unique():
            for (task, dataset) in aggregated_data.groupby(['task', 'dataset']).groups.keys():
                plot_file = file_suffix.format(task=task, dataset=dataset, metric=metric).replace(" ", "_")
                plot_path = os.path.join(plot_dir, plot_file)
                if os.path.exists(plot_path):
                    img = Image(plot_path)
                    img.width = image_width
                    img.height = image_height
                    img.anchor = f"{get_column_letter(col_offset)}{start_row}"  # Place image in the correct column
                    ws.add_image(img)
                    col_offset += 6  # Adjust column spacing for images

        start_row += 15  # Move to the next row for the next category

    # Calculate starting position for sequential categories
    sequential_start_row = start_row + 5  # Leave extra space after parallel layout
    sequential_start_column = rightmost_column + 10  # Shift sequential plots to the right

    # Add sequential layout plots
    for category, file_suffix in sequential_categories.items():
        # Add category name
        ws.cell(row=sequential_start_row, column=sequential_start_column, value=category)
        sequential_start_row += 2  # Leave a small gap for better readability

        plot_path = os.path.join(plot_dir, file_suffix)
        if os.path.exists(plot_path):
            img = Image(plot_path)
            img.width = image_width
            img.height = image_height
            img.anchor = f"{get_column_letter(sequential_start_column)}{sequential_start_row}"  # Place plot to the right
            ws.add_image(img)
            sequential_start_row += 22  # Adjust row spacing for resized images

    # Save the Excel file
    wb.save(output_file)
    print(f"Excel file saved to {output_file}")
