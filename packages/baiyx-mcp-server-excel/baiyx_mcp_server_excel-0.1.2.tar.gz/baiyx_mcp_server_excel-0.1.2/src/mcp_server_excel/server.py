import os
from typing import List, Optional, Dict, Any
import pandas as pd
from openpyxl import load_workbook, Workbook
import importlib.metadata
from openpyxl.utils import get_column_letter, column_index_from_string

from mcp.server.fastmcp.server import FastMCP
from mcp.shared.exceptions import McpError
from pydantic import BaseModel


class SheetInfo(BaseModel):
    """Excel工作表信息模型
    
    实际示例:
    ```python
    sheet_info = SheetInfo(
        name="销售数据",
        headers=["日期", "产品", "数量", "单价", "总价"],
        data_range="A1:E100",
        row_count=100,
        column_count=5
    )
    ```

    完整数据示例:
    ```json
    {
        "name": "销售数据",
        "headers": ["日期", "产品", "数量", "单价", "总价"],
        "data_range": "A1:E100",
        "row_count": 100,
        "column_count": 5
    }
    ```
    """
    name: str  # 工作表名称
    headers: List[str]  # 表头列表（第一行的数据）
    data_range: str  # 数据范围（例如："A1:E100"）
    row_count: int  # 总行数
    column_count: int  # 总列数


class WorkbookInfo(BaseModel):
    """Excel工作簿信息模型
    
    实际示例:
    ```python
    workbook_info = WorkbookInfo(
        file_path="sales_2024.xlsx",
        sheets=[
            SheetInfo(
                name="Q1销售",
                headers=["日期", "产品", "数量", "单价", "总价"],
                data_range="A1:E100",
                row_count=100,
                column_count=5
            ),
            SheetInfo(
                name="Q2销售",
                headers=["日期", "产品", "数量", "单价", "总价"],
                data_range="A1:E150",
                row_count=150,
                column_count=5
            )
        ],
        version="0.1.0"
    )
    ```

    完整数据示例:
    ```json
    {
        "file_path": "sales_2024.xlsx",
        "sheets": [
            {
                "name": "Q1销售",
                "headers": ["日期", "产品", "数量", "单价", "总价"],
                "data_range": "A1:E100",
                "row_count": 100,
                "column_count": 5
            },
            {
                "name": "Q2销售",
                "headers": ["日期", "产品", "数量", "单价", "总价"],
                "data_range": "A1:E150",
                "row_count": 150,
                "column_count": 5
            }
        ],
        "version": "0.1.0"
    }
    ```
    """
    file_path: str  # Excel文件路径
    sheets: List[SheetInfo]  # 工作表信息列表
    version: str  # 服务器版本号


class ExcelData(BaseModel):
    """Excel数据模型
    
    实际示例:
    ```python
    sales_data = ExcelData(
        data=[
            ["日期", "产品", "数量", "单价", "总价"],
            ["2024-03-01", "笔记本电脑", 5, 6999.00, "=C2*D2"],
            ["2024-03-01", "显示器", 10, 1999.00, "=C3*D3"],
            ["2024-03-02", "键盘", 20, 299.00, "=C4*D4"],
            ["2024-03-02", "鼠标", 30, 199.00, "=C5*D5"]
        ],
        range="A1:E5"
    )
    ```

    完整数据示例:
    ```json
    {
        "data": [
            ["日期", "产品", "数量", "单价", "总价"],
            ["2024-03-01", "笔记本电脑", 5, 6999.00, 34995.00],
            ["2024-03-01", "显示器", 10, 1999.00, 19990.00],
            ["2024-03-02", "键盘", 20, 299.00, 5980.00],
            ["2024-03-02", "鼠标", 30, 199.00, 5970.00]
        ],
        "range": "A1:E5"
    }
    ```
    """
    data: List[List[Any]]  # 二维数组形式的数据
    range: str  # 数据范围


class ExcelFormula(BaseModel):
    """Excel公式模型
    
    实际示例:
    ```python
    sales_formulas = ExcelFormula(
        formulas=[
            "=SUM(E2:E5)",              # 总销售额
            "=AVERAGE(E2:E5)",          # 平均订单金额
            "=COUNTIF(B2:B5,\"键盘\")",  # 键盘销售次数
            "=SUMIF(B2:B5,\"显示器\",E2:E5)",  # 显示器销售总额
            "=MAX(E2:E5)",              # 最大订单金额
            "=MIN(E2:E5)"               # 最小订单金额
        ],
        range="F2:F7"
    )
    ```

    完整数据示例:
    ```json
    {
        "formulas": [
            "=SUM(E2:E5)",
            "=AVERAGE(E2:E5)",
            "=COUNTIF(B2:B5,\"键盘\")",
            "=SUMIF(B2:B5,\"显示器\",E2:E5)",
            "=MAX(E2:E5)",
            "=MIN(E2:E5)"
        ],
        "range": "F2:F7"
    }
    ```

    常用公式示例:
    1. 数学运算
       - 基础运算: "=A1+B1", "=C1-D1", "=E1*F1", "=G1/H1"
       - 幂运算: "=A1^2", "=B1^3"
       - 四舍五入: "=ROUND(A1,2)", "=ROUNDUP(B1,0)", "=ROUNDDOWN(C1,1)"
       
    2. 统计函数
       - 求和: "=SUM(A1:A10)", "=SUMIF(A1:A10,\">100\")"
       - 平均值: "=AVERAGE(B1:B10)", "=AVERAGEIF(A1:A10,\">0\",B1:B10)"
       - 计数: "=COUNT(C1:C10)", "=COUNTIF(C1:C10,\">=1000\")"
       - 最值: "=MAX(D1:D10)", "=MIN(D1:D10)"
       
    3. 文本函数
       - 拼接: "=CONCATENATE(A1,\" \",B1)"
       - 截取: "=LEFT(C1,2)", "=RIGHT(D1,3)", "=MID(E1,2,3)"
       - 大小写: "=UPPER(F1)", "=LOWER(G1)", "=PROPER(H1)"
       
    4. 日期函数
       - 当前日期: "=TODAY()", "=NOW()"
       - 提取: "=YEAR(A1)", "=MONTH(B1)", "=DAY(C1)"
       - 计算: "=DATEDIF(A1,B1,\"Y\")", "=NETWORKDAYS(C1,D1)"
       
    5. 逻辑函数
       - 条件: "=IF(A1>100,\"高\",\"低\")"
       - 多条件: "=IF(AND(A1>0,A1<100),\"正常\",\"异常\")"
       - 嵌套: "=IF(A1>100,\"高\",IF(A1>50,\"中\",\"低\"))"
       
    6. 查找函数
       - 纵向查找: "=VLOOKUP(A1,B1:D10,2,FALSE)"
       - 横向查找: "=HLOOKUP(A1,B1:B10,2,FALSE)"
       - 匹配: "=MATCH(A1,B1:B10,0)"
    """
    formulas: List[str]  # 公式列表
    range: str  # 公式范围


class ExcelServer:
    def __init__(self):
        self.mcp = FastMCP("mcp-excel")
        self._tools = {}
        self.setup_tools()
        
        version = importlib.metadata.version("baiyx-mcp-server-excel")
        print(f"\n📊 MCP Excel Server v{version} starting...")
        print("✨ Server is ready to handle requests!\n")

    def setup_tools(self):
        def get_workbook_info(file_path: str) -> WorkbookInfo:
            """获取Excel工作簿的详细信息，包括所有工作表的信息
            
            参数:
                file_path (str): Excel文件路径，例如："sales_2024.xlsx"
            
            返回:
                WorkbookInfo: 工作簿信息对象
                
            实际使用示例:
            ```python
            # 读取销售数据文件信息
            workbook = get_workbook_info("sales_2024.xlsx")
            
            # 打印工作簿信息
            print(f"文件路径: {workbook.file_path}")
            print(f"工作表数量: {len(workbook.sheets)}")
            
            # 遍历所有工作表
            for sheet in workbook.sheets:
                print(f"\n工作表: {sheet.name}")
                print(f"表头: {sheet.headers}")
                print(f"数据范围: {sheet.data_range}")
                print(f"行数: {sheet.row_count}")
                print(f"列数: {sheet.column_count}")
            ```

            输出示例:
            ```
            文件路径: sales_2024.xlsx
            工作表数量: 2

            工作表: Q1销售
            表头: ['日期', '产品', '数量', '单价', '总价']
            数据范围: A1:E100
            行数: 100
            列数: 5

            工作表: Q2销售
            表头: ['日期', '产品', '数量', '单价', '总价']
            数据范围: A1:E150
            行数: 150
            列数: 5
            ```
            """
            if not os.path.exists(file_path):
                raise McpError(f"File not found: {file_path}")
            
            wb = load_workbook(file_path, data_only=False)
            sheets_info = []
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                data_rows = list(ws.rows)
                if not data_rows:
                    continue
                    
                headers = [str(cell.value) for cell in data_rows[0] if cell.value is not None]
                if not headers:
                    continue
                    
                max_row = ws.max_row
                max_col = ws.max_column
                data_range = f"A1:{chr(64 + max_col)}{max_row}"
                
                sheets_info.append(SheetInfo(
                    name=sheet_name,
                    headers=headers,
                    data_range=data_range,
                    row_count=max_row,
                    column_count=max_col
                ))
            
            return WorkbookInfo(
                file_path=file_path,
                sheets=sheets_info,
                version=importlib.metadata.version("baiyx-mcp-server-excel")
            )

        def read_sheet_data(file_path: str, sheet_name: str, range: Optional[str] = None) -> ExcelData:
            """读取Excel工作表中的数据
            
            参数:
                file_path (str): Excel文件路径，例如："sales_2024.xlsx"
                sheet_name (str): 工作表名称，例如："Q1销售"
                range (str, optional): 要读取的单元格范围，例如："A1:E5"
            
            返回:
                ExcelData: 包含数据和范围信息的对象
                
            实际使用示例:
            ```python
            # 读取指定范围的销售数据
            sales_data = read_sheet_data("sales_2024.xlsx", "Q1销售", "A1:E5")
            
            # 打印数据
            print("销售数据:")
            for row in sales_data.data:
                print(row)
            print(f"数据范围: {sales_data.range}")
            
            # 读取整个工作表
            all_data = read_sheet_data("sales_2024.xlsx", "Q1销售")
            print(f"\n总行数: {len(all_data.data)}")
            print(f"总列数: {len(all_data.data[0])}")
            ```

            输出示例:
            ```
            销售数据:
            ['日期', '产品', '数量', '单价', '总价']
            ['2024-03-01', '笔记本电脑', 5, 6999.00, 34995.00]
            ['2024-03-01', '显示器', 10, 1999.00, 19990.00]
            ['2024-03-02', '键盘', 20, 299.00, 5980.00]
            ['2024-03-02', '鼠标', 30, 199.00, 5970.00]
            数据范围: A1:E5

            总行数: 100
            总列数: 5
            ```
            """
            if not os.path.exists(file_path):
                raise McpError(f"File not found: {file_path}")
            
            df = pd.read_excel(file_path, sheet_name=sheet_name)
            
            if range:
                # Parse range (e.g., "A1:C10")
                start_col, start_row, end_col, end_row = self._parse_range(range)
                # Convert to 0-based index
                start_row -= 1
                end_row -= 1
                start_col = ord(start_col) - ord('A')
                end_col = ord(end_col) - ord('A') + 1
                
                df = df.iloc[start_row:end_row, start_col:end_col]
            
            return ExcelData(
                data=df.values.tolist(),
                range=range or f"A1:{chr(64 + len(df.columns))}{len(df) + 1}"
            )

        def read_sheet_formula(file_path: str, sheet_name: str, cell_range: Optional[str] = None) -> ExcelFormula:
            """读取Excel工作表中的公式
            
            参数:
                file_path (str): Excel文件路径，例如："sales_2024.xlsx"
                sheet_name (str): 工作表名称，例如："Q1销售"
                cell_range (str, optional): 要读取的单元格范围，例如："F2:F7"
            
            返回:
                ExcelFormula: 包含公式列表和范围信息的对象
                
            实际使用示例:
            ```python
            # 读取销售统计公式
            formulas = read_sheet_formula("sales_2024.xlsx", "Q1销售", "F2:F7")
            
            # 打印公式
            print("销售统计公式:")
            for formula in formulas.formulas:
                print(formula)
            print(f"公式范围: {formulas.range}")
            
            # 读取所有公式
            all_formulas = read_sheet_formula("sales_2024.xlsx", "Q1销售")
            print(f"\n总公式数: {len(all_formulas.formulas)}")
            ```

            输出示例:
            ```
            销售统计公式:
            =SUM(E2:E5)              # 总销售额: 66935.00
            =AVERAGE(E2:E5)          # 平均订单金额: 16733.75
            =COUNTIF(B2:B5,"键盘")   # 键盘销售次数: 1
            =SUMIF(B2:B5,"显示器",E2:E5)  # 显示器销售总额: 19990.00
            =MAX(E2:E5)              # 最大订单金额: 34995.00
            =MIN(E2:E5)              # 最小订单金额: 5970.00
            公式范围: F2:F7

            总公式数: 6
            ```
            """
            if not os.path.exists(file_path):
                raise McpError(f"File not found: {file_path}")
            
            wb = load_workbook(file_path, data_only=False)
            ws = wb[sheet_name]
            
            formulas = []
            if cell_range:
                start_col, start_row, end_col, end_row = self._parse_range(cell_range)
                for row_idx in range(int(start_row), int(end_row) + 1):
                    row_formulas = []
                    for col_idx in range(ord(start_col) - ord('A'), ord(end_col) - ord('A') + 1):
                        cell = ws[f"{chr(65 + col_idx)}{row_idx}"]
                        row_formulas.append(cell.value if cell.value and str(cell.value).startswith('=') else '')
                    formulas.append(row_formulas)
            else:
                cell_range = f"A1:{chr(64 + ws.max_column)}{ws.max_row}"
                for row in ws.iter_rows():
                    row_formulas = [cell.value if cell.value and str(cell.value).startswith('=') else '' for cell in row]
                    formulas.append(row_formulas)
            
            return ExcelFormula(
                formulas=[formula for row in formulas for formula in row if formula],
                range=cell_range
            )

        def write_sheet_data(file_path: str, sheet_name: str, start_cell: str, data: List[List[Any]]) -> bool:
            """写入数据到Excel工作表
            
            参数:
                file_path (str): Excel文件路径，例如："sales_2024.xlsx"
                sheet_name (str): 工作表名称，例如："Q1销售"
                start_cell (str): 起始单元格，例如："A1"
                data (List[List[Any]]): 要写入的二维数组数据
            
            返回:
                bool: 写入成功返回True，失败抛出异常
                
            实际使用示例:
            ```python
            # 准备销售数据
            sales_data = [
                ["日期", "产品", "数量", "单价", "总价"],
                ["2024-03-01", "笔记本电脑", 5, 6999.00, "=C2*D2"],
                ["2024-03-01", "显示器", 10, 1999.00, "=C3*D3"],
                ["2024-03-02", "键盘", 20, 299.00, "=C4*D4"],
                ["2024-03-02", "鼠标", 30, 199.00, "=C5*D5"]
            ]
            
            # 写入数据
            success = write_sheet_data("sales_2024.xlsx", "Q1销售", "A1", sales_data)
            
            if success:
                print("数据写入成功！")
                print("写入的数据:")
                for row in sales_data:
                    print(row)
            ```

            输出示例:
            ```
            数据写入成功！
            写入的数据:
            ['日期', '产品', '数量', '单价', '总价']
            ['2024-03-01', '笔记本电脑', 5, 6999.00, '=C2*D2']
            ['2024-03-01', '显示器', 10, 1999.00, '=C3*D3']
            ['2024-03-02', '键盘', 20, 299.00, '=C4*D4']
            ['2024-03-02', '鼠标', 30, 199.00, '=C5*D5']
            ```
            """
            try:
                try:
                    wb = load_workbook(filename=file_path)
                except FileNotFoundError:
                    wb = Workbook()
                
                if sheet_name not in wb.sheetnames:
                    wb.create_sheet(sheet_name)
                ws = wb[sheet_name]
                
                # 解析起始单元格
                start_col = ''.join(c for c in start_cell if c.isalpha())
                start_row = int(''.join(c for c in start_cell if c.isdigit()))
                
                # 写入数据
                for i, row_data in enumerate(data):
                    for j, value in enumerate(row_data):
                        col = get_column_letter(column_index_from_string(start_col) + j)
                        ws[f"{col}{start_row + i}"] = value
                
                wb.save(file_path)
                return True
                
            except Exception as e:
                raise McpError(str(e))

        def write_sheet_formula(file_path: str, sheet_name: str, range: str, formulas: List[str]) -> bool:
            """写入公式到Excel工作表
            
            参数:
                file_path (str): Excel文件路径，例如："sales_2024.xlsx"
                sheet_name (str): 工作表名称，例如："Q1销售"
                range (str): 起始单元格，例如："F2"
                formulas (List[str]): 要写入的公式列表
            
            返回:
                bool: 写入成功返回True，失败抛出异常
                
            实际使用示例:
            ```python
            # 准备统计公式
            stat_formulas = [
                "=SUM(E2:E5)",              # 总销售额
                "=AVERAGE(E2:E5)",          # 平均订单金额
                "=COUNTIF(B2:B5,\"键盘\")",  # 键盘销售次数
                "=SUMIF(B2:B5,\"显示器\",E2:E5)",  # 显示器销售总额
                "=MAX(E2:E5)",              # 最大订单金额
                "=MIN(E2:E5)"               # 最小订单金额
            ]
            
            # 写入公式
            success = write_sheet_formula("sales_2024.xlsx", "Q1销售", "F2", stat_formulas)
            
            if success:
                print("公式写入成功！")
                print("写入的公式:")
                for formula in stat_formulas:
                    print(formula)
            ```

            输出示例:
            ```
            公式写入成功！
            写入的公式:
            =SUM(E2:E5)              # 总销售额
            =AVERAGE(E2:E5)          # 平均订单金额
            =COUNTIF(B2:B5,"键盘")   # 键盘销售次数
            =SUMIF(B2:B5,"显示器",E2:E5)  # 显示器销售总额
            =MAX(E2:E5)              # 最大订单金额
            =MIN(E2:E5)              # 最小订单金额
            ```
            """
            try:
                wb = load_workbook(file_path) if os.path.exists(file_path) else None
                if wb is None:
                    wb = Workbook()
                    wb.create_sheet(sheet_name)
                elif sheet_name not in wb.sheetnames:
                    wb.create_sheet(sheet_name)
                
                ws = wb[sheet_name]
                start_col, start_row, _, _ = self._parse_range(range)
                
                for i, formula in enumerate(formulas):
                    col = chr(ord(start_col) + (i % ws.max_column))
                    row = start_row + (i // ws.max_column)
                    cell = ws[f"{col}{row}"]
                    cell.value = formula
                
                wb.save(file_path)
                return True
            except Exception as e:
                raise McpError(message=f"写入公式失败: {str(e)}")

        # 注册工具函数
        self._tools = {
            'get_workbook_info': get_workbook_info,
            'read_sheet_data': read_sheet_data,
            'read_sheet_formula': read_sheet_formula,
            'write_sheet_data': write_sheet_data,
            'write_sheet_formula': write_sheet_formula
        }

        # 注册到 MCP
        for name, func in self._tools.items():
            self.mcp.tool()(func)

        return self._tools

    def _parse_range(self, range_str: str) -> tuple[str, int, str, int]:
        """Parse Excel range string (e.g., 'A1:C10') into components"""
        try:
            start, end = range_str.split(':')
            start_col = ''.join(c for c in start if c.isalpha())
            start_row = int(''.join(c for c in start if c.isdigit()))
            end_col = ''.join(c for c in end if c.isalpha())
            end_row = int(''.join(c for c in end if c.isdigit()))
            return start_col, start_row, end_col, end_row
        except Exception as e:
            raise McpError(f"Invalid range format. Expected 'A1:C10', got '{range_str}': {str(e)}")


async def serve() -> None:
    server = ExcelServer()
    await server.mcp.run_stdio_async() 