import hashlib
import json
import math
import openpyxl
import os
import pymssql
import pymysql
import re
import requests
import sqlalchemy
import time
import zipfile

import pandas as pd
import numpy as np

from openpyxl.utils import get_column_letter
from tqdm import tqdm


class GetEcData:

    def __init__(self, user_name, user_token):
        self.wsdl = 'http://openapi-web.eccang.com/openApi/api/unity'
        self.headers = {
            'ContentType': 'application/json'
        }

        self.user_name = user_name
        self.user_token = user_token

    def __concat_params(self, biz_content: dict, interface_method: str):
        """
        处理post data, 生成key
        :param biz_content:  详细调用参数
        :param interface_method:  易仓的信息 getStockOrderList
        :return:
        """
        post_data = {
            "app_key": self.user_name,
            "biz_content": json.dumps(biz_content),
            "charset": "UTF-8",
            "interface_method": interface_method,
            "nonce_str": "113456",
            "service_id": "E7HPYV",
            "sign_type": "MD5",
            "timestamp": int(time.time() * 1000),
            "version": "v1.0.0"
        }

        # 将字典转化为易仓需要的加密形式
        post_data_str = ''
        for one_key, one_value in zip(post_data.keys(), post_data.values()):
            if type(one_value) == dict:
                one_value = json.dumps(one_value).replace(': ', ':')
            post_data_str += one_key
            post_data_str += '='
            post_data_str += str(one_value)
            post_data_str += '&'

        post_data_str = post_data_str[:-1]
        post_data_str += self.user_token

        # 对组合后的信息进行加密md5
        post_data['sign'] = hashlib.md5(bytes(post_data_str, encoding='utf-8')).hexdigest()

        return post_data

    def __get_data(self, biz_content: dict, interface_method: str, key_word: str = 'data'):
        """
        获取单页数据，把传入的参数转成json格式，向api请求，提取response.text里的data数据
        :param biz_content:
        :param interface_method:
        :param key_word: 要返回的关键词
        :return:
        """
        concated_params = self.__concat_params(biz_content, interface_method)

        # 获取response
        res = requests.post(self.wsdl, json=concated_params, headers=self.headers)
        # 把response的text的json格式转换成字典格式
        try:
            page_info = json.loads(res.text)
            # 打印异常信息
            if page_info['message'] not in ['Success', 'ok']:
                print(page_info['message'])  
                print(f'当前请求的biz_content：{biz_content}')
                print(f'当前请求的interface_method：{interface_method}')
        except:
            page_info = {'message': '系统异常'}
            print(res.text)

        # 判断是否超时
        try:
            # 根据传入的键返回值
            page_data = json.loads(page_info.get('biz_content')).get(key_word)
        except:
            print('系统异常，易仓可能超时')
            print(f'当前请求的biz_content：{biz_content}')
            print(f'当前请求的interface_method：{interface_method}')
            print(page_info)
            page_data = ''

        return page_data

    def get_data(self, biz_content: dict, interface_method: str, special_param: str = None):
        """
        https://open.eccang.com/#/documentCenter?docId=1287&catId=0-225-225,0-177
        获取请求的数据
        :param biz_content:
        :param interface_method:
        :param special_param: 特殊参数，传入该参数后不会尝试获取数据的最大行数，而是遍历biz_content中的该参数列表
        :return:
        """
        # 0 参数设置
        # 默认页数
        if not biz_content.get('page_size'):
            biz_content['page_size'] = 20

        list_df = []
        if not special_param:
            # 1 获取最大页数
            record_rows = self.__get_data(biz_content, interface_method, 'total_count')
            if not record_rows:
                record_rows = self.__get_data(biz_content, interface_method, 'total')
            # 向上取整
            max_page = math.ceil(int(record_rows) / biz_content.get('page_size'))

            # 2 按页获取数据
            print('按页获取数据')
            for i in tqdm(range(1, max_page + 1)):
                time.sleep(5)  # 易仓限制了请求频率，只能自己减少了
                # 2.1 调整键值对
                biz_content['page'] = i
                # 2.2 获取对应页数的数据
                pg_data = self.__get_data(biz_content, interface_method)
                if pg_data:
                    list_df.append(pd.DataFrame(pg_data))
        else:
            list_param = biz_content[special_param]  # 参数列表，比如订单号
            lens = len(list_param)
            # 1 遍历special_param，每次1个
            print(f'根据{special_param}，每次获取1个数据')
            for i in tqdm(range(0, lens, 1)):
                # 2.2 获取对应页数的数据
                biz_content[special_param] = list_param[i: i + 1]
                pg_data = self.__get_data(biz_content, interface_method)
                if pg_data:
                    list_df.append(pd.DataFrame(pg_data))

        return list_df


class Mysql:

    def __init__(self, host, user, password, database):
        self.conn = pymysql.connect(host=host, user=user, password=password, database=database, charset='utf8')
        self.cur = self.conn.cursor()

    def exec_query(self, sql):
        self.cur.execute(sql)
        self.conn.commit()
        result = self.cur.fetchall()
        for row in result:
            print(row)

    def close(self):
        self.conn.close()


def any_files(folder_path: str) -> bool:
    """
    给文件地址，判断里面有没有文件，有的话返回True，反之False
    :param folder_path:
    :return:
    """
    for root, dirs, files in os.walk(folder_path):
        if files:
            return True
    return False


def excel_process(file_path):
    """
    调整列宽，冻结首行，添加筛选
    freeze title, adjust width of columns, open filter
    :param file_path: path of file
    :return:
    """
    print('调整列宽，冻结首行，添加筛选')
    # 修改下述参数即可使用，Excel名称及Sheet名称即可
    work_book = openpyxl.load_workbook(file_path)
    for sheet in work_book.sheetnames:
        work_book[sheet].freeze_panes = 'A2'
        work_sheet = work_book[sheet]
        # 设置一个字典用于保存列宽数据
        dim_cols = {}
        # 遍历表格数据，获取自适应列宽数据
        for row in work_sheet.rows:
            for cell in row:
                if cell.value:
                    # 遍历整个表格，把该列所有的单元格文本进行长度对比，找出最长的单元格
                    # 在对比单元格文本时需要将中文字符识别为1.7个长度，英文字符识别为1个，这里只需要将文本长度直接加上中文字符数量即可
                    # re.findall('([\u4e00-\u9fa5])', cell.value)能够识别大部分中文字符
                    cell_len = 0.5 * len(re.findall('([\u4e00-\u9fa5])', str(cell.value))) + len(str(cell.value))
                    dim_cols[cell.column] = max((dim_cols.get(cell.column, 0), cell_len))
        for col, value in dim_cols.items():
            # 设置列宽，get_column_letter用于获取数字列号对应的字母列号，最后值+2是用来调整最终效果的，限制最小宽度10， 最大宽度为30
            if value > 28:
                work_sheet.column_dim_colensions[get_column_letter(col)].width = 30
            elif value < 8:
                work_sheet.column_dim_colensions[get_column_letter(col)].width = 10
            else:
                work_sheet.column_dim_colensions[get_column_letter(col)].width = value + 2
        dict_num_to_alphabet = {
            1: 'A', 2: 'B', 3: 'C', 4: 'D', 5: 'E', 6: 'F', 7: 'G', 8: 'H', 9: 'I', 10: 'J',
            11: 'K', 12: 'L', 13: 'M', 14: 'N', 15: 'O', 16: 'P', 17: 'Q', 18: 'R', 19: 'S', 20: 'T',
            21: 'U', 22: 'V', 23: 'W', 24: 'X', 25: 'Y', 26: 'Z',
        }
        # 获取第一行
        rows = work_sheet.iter_rows(max_row=1, values_only=True)
        max_col = 0
        for cell in rows:
            max_col = len(cell)
        # 对第一行添加过滤功能
        filters = work_sheet.auto_filter
        filters.ref = 'A:' + dict_num_to_alphabet.get(max_col)
    work_book.save(file_path)
    print('end')
    print('$' * 20)


def incremental_update(
        df: pd.DataFrame,
        table_name: str,
        merge_cols: list,
        conn: sqlalchemy.engine,
        server,
        user,
        password,
        database,
        dtype: dict = None,
        calculate_cols=None,
        dim_col: str = None,
        select = '',
):
    """
    增量更新df至数据库的table_name中
    :param df: 需要被传入的DataFrame
    :param table_name: 需要更新的数据库表名
    :param merge_cols: 匹配的字段列表
    :param conn: 数据库连接
    :param server: 服务器地址
    :param user: 用户名
    :param password: 密码
    :param database: db名
    :param dtype: 特殊字段类型，默认空
    :param calculate_cols: 需要计算的字段，默认空，可接受字符串（单列）、列表（多列）
    :param dim_col: 维度列表，如果存在此参数，则不会根据merge_cols来判断本地与数据库的重复数据，而是通过此参数判断
    :param select: 选项，默认空
    """
    # dtype设置为空白字典
    if dtype is None:
        dtype = {}

    # 获取当前系统用户名
    user_program = os.getlogin()
    # 重置索引，不然匹配时
    df = df.reset_index(drop=True)
    # 检测数据库中重复记录，有计算列的话会取计算列
    sql = f"select {', '.join(merge_cols)}" + ", id" if merge_cols else f"select {dim_col}" + ", id"
    if calculate_cols:
        if type(calculate_cols) == str:
            sql += f', {calculate_cols}'
        elif type(calculate_cols) == list:
            sql = sql + ', ' + ', '.join(calculate_cols)
    sql += f' from {table_name}'
    df_db = pd.read_sql(sql, conn, dtype=dtype)

    # 本地文件和数据库都有的数据
    if merge_cols:
        df_inner = df_db.merge(df, 'inner', merge_cols, )
    else:
        dim_tuple = tuple(df[dim_col].drop_duplicates().astype('str').values)  # 本地的dim_col去重后的维度元组
        df_inner = df_db.loc[lambda d: d[dim_col].isin(dim_tuple)]

    # 只有本地文件有的数据
    if merge_cols:
        df_not_inner = df.merge(df_inner[merge_cols].assign(mark=1), 'left', merge_cols, ).loc[
            lambda d: d['mark'].isnull()].drop('mark', axis=1)
    else:
        df_not_inner = pd.DataFrame({})

    nums_db = df_db.shape[0]  # 数据库的总数量
    nums_df = df.shape[0]  # 本地的总数量
    nums_inner = df_inner.shape[0]  # 两者交集的数量
    nums_only_df = df_not_inner.shape[0]  # 非交集的数量
    print('-' * 50)
    print(f'''数据库中“{table_name}”数据：{nums_db}条''')
    if dim_col:
        # 数据库的对应维度数量
        nums_db_dim = df_db.loc[lambda d: d[dim_col].isin(dim_tuple)].shape[0]
        print(f'''\t其中维度：{dim_col}：{dim_tuple}的：{nums_db_dim}条。''')
    print('-' * 50)
    print(f'''本地数据：{nums_df}条''')
    print('其中：') if nums_inner > 0 else None
    if nums_inner > 0:
        if dim_col:
            # 本地的对应维度数量
            nums_df_dim = df.loc[lambda d: d[dim_col].isin(dim_tuple)].shape[0]
            print(f'''\t其中维度：{dim_col}：{dim_tuple}的：{nums_df_dim}条。''')
        else:
            print(f'''\t两者都有的：{nums_inner}条''')
    print(f'''\t本地新的：{nums_only_df}条''') if nums_only_df > 0 else None

    # 有计算列的额外提示语句
    add_info = ''
    c1, c2 = '', ''
    if calculate_cols is not None:
        if type(calculate_cols) == str:
            # c1是数据库的金额，c2是本地文件的金额，如果有比较列，生成有差异的df
            c1, c2 = calculate_cols + '_x', calculate_cols + '_y'
            df_diff = df_db.merge(df, 'outer', merge_cols).loc[lambda d: (d[c1] != d[c2]) & (d[c2].notnull())]
            num_diff = df_diff.shape[0] - nums_inner
            add_info += f'两者都有的但“{calculate_cols}”数值不一致的：{num_diff}条' if num_diff > 0 else ''
        elif type(calculate_cols) == list:
            for calculate_col in calculate_cols:
                c1, c2 = calculate_col + '_x', calculate_col + '_y'
                df_diff = df_db.merge(df, 'outer', merge_cols).loc[lambda d: (d[c1] != d[c2]) & (d[c2].notnull())]
                num_diff = df_diff.shape[0]
                add_info += f'两者都有的但“{calculate_col}”数值不一致的：{num_diff}条\n\t' if num_diff > 0 else ''
    print(f'''\t{add_info}''') if add_info else None
    print('-' * 50)

    # 打印出重复信息
    if nums_inner > 5:
        df_print = df_inner.sample(1).reset_index()
    elif nums_inner == 0:
        df_print = pd.DataFrame({})
    else:
        df_print = df_inner.reset_index()
    for index in df_print.index:
        print('两者都有的随意一条数据：')
        print(df_print.loc[index])
        print('-' * 50)

    info = ('请输入指令以继续（直接回车可跳过）：'
            '\n1：清空数据库，再导入本地数据（此操作会清空历史数据，请慎重选择！）'
            '\n2: 只上传非重复记录（只添加新的）'
            '\n3: 删除数据不一致的，再上传非重复记录（更新老的，添加新的）'
            '\n\n')

    if add_info:
        info += '本地的数据跟数据库的相比，数值有变化，建议选：3'
    else:
        if nums_df == nums_inner:
            info += '本地的数据在数据库都有，建议回车跳过'
        elif nums_df == nums_only_df:
            info += '本地的数据在数据库都没有，建议选：2'
        else:
            info += '本地的数据有部分和数据库重合，建议选：3'
    info += '\n'

    # 如果传入的select有值，就用传入的，没有就用提示的
    if not select:
        select = input(info)

    # 要删除的id
    ids_to_be_deleted = df_inner['id'].to_list() if select == '3' else []
    # 要导入的数据，1和3都要导入所有本地数据
    df_upload = df_not_inner if select == '2' else df
    mysql = Mysql(server, user, password, database)

    # 先删除
    if select == '1':
        print(f'清空{table_name}')
        mysql.exec_query(f'truncate table {table_name}')
    elif select == '3':
        ids_len = len(ids_to_be_deleted)
        if ids_len > 0:
            # 删掉数据库中这些数据
            print(f'删除{table_name}中{ids_len}条记录')
            # 有维度的话按照维度删
            if dim_col:
                # 如果元组只有1个元素，把它变成字符串后替换掉逗号
                if len(dim_tuple) == 1:
                    dim_tuple = str(dim_tuple).replace(',', '')
                mysql.exec_query(f'delete {table_name} where {dim_col} in {dim_tuple}')
            else:
                for i in tqdm(range(0, ids_len, 1000)):
                    mysql.exec_query(f'delete from {table_name} where id in {tuple(ids_to_be_deleted[i: i + 1000])}')
            print(f'\n{user_program}删除了{table_name}的{ids_len}条记录')

    if select != '':
        # 导入数据库
        upload_records = df_upload.shape[0]
        # 再重置一次索引
        df_upload = df_upload.reset_index(drop=True)

        # 尝试添加update_time
        columns_db = pd.read_sql(f'select * from {table_name} limit 1', conn).columns.to_list()
        if 'update_time' in columns_db:
            print('添加update_time')
            df_upload['update_time'] = pd.Timestamp.now()

        print(f'{upload_records}条记录等待被导入至{table_name}')
        for i in tqdm(range(0, upload_records, 1000)):
            df_upload.loc[i: i + 999].to_sql(f'{table_name}', conn, index=False, if_exists='append')
        print(f'{user_program}导入了{upload_records}条记录至{table_name}')
    else:
        print('跳过')


def lambda_f(n: str, sep: str = ','):
    """
    根据输入的关键词返回对应的匿名函数
    :param n: lam_multi_to_unique_single, s: lam_multi_to_single
    :param sep: 分隔符，默认逗号
    :return lambda function
    """
    lam_multi_to_unique_single = lambda x: re.sub('^,|,$', '', sep.join(set(x)))  # 多行转唯一一行
    lam_multi_to_single = lambda x: re.sub('^,|,$', '', sep.join(list(x)))  # 多行转一行
    return {'us': lam_multi_to_unique_single, 's': lam_multi_to_single}.get(n)


def print_date_ranges(series):
    # 确保Series为datetime类型，并按日期排序
    s = series.sort_values().reset_index(drop=True)

    # 计算日期差是否超过1天，标记新分组的起点
    mask = (s.diff() > pd.Timedelta(days=1)).fillna(False)
    group_ids = mask.cumsum()

    # 按分组聚合，获取每组的最小和最大日期
    groups = s.groupby(group_ids).agg([('start', 'min'), ('end', 'max')])

    # 生成结果字符串
    result = []
    for _, row in groups.iterrows():
        start_str = row['start'].strftime('%Y-%m-%d')
        end_str = row['end'].strftime('%Y-%m-%d')
        if start_str == end_str:
            result.append(f"'{start_str}'")
        else:
            result.append(f"'{start_str}'至'{end_str}'")

    # 拼接最终输出
    print('数据库现有的日期范围：')
    print(', '.join(result))


def input_date_period():
    """创建日期范围，默认昨天，返回开始日期、结束日期、账单日期，账单日期为开始日期/结束日期"""
    # 1 输入日期
    default_month = (pd.Timestamp.now().to_period('D') - 1).strftime('%Y-%m-%d')
    start_date = input(f'请输入开始日期，格式：yyyy-mm-dd，如果只输入yyyy-mm则会下载整月数据，回车则输入该日：\n{default_month}\n')
    if not start_date:
        start_date = default_month
    # 检测是整月还是日期
    if re.search('^\d{4}-\d{1,2}$', start_date):
        start_date = pd.to_datetime(start_date, format='%Y-%m').date()
        end_date = (start_date + pd.offsets.MonthEnd(0)).date()
        print(f'您输入的为整月，日期范围为：{start_date} 至 {end_date}')
    else:
        start_date = pd.to_datetime(start_date, format='%Y-%m-%d').date()
        end_date = input(f'请输入结束日期，格式：yyyy-mm-dd，直接回车的话结束日期为：{str(start_date)}\n')
        if end_date == '':
            end_date = start_date
        else:
            end_date = pd.to_datetime(end_date, format='%Y-%m-%d').date()
    start_date = str(start_date)
    end_date = str(end_date)
    bill_date = start_date + '/' + end_date
    print(f'您选择的日期范围为：{start_date}至{end_date}')

    return start_date, end_date, bill_date


def rename_dict(conn, db_table_name, table_name, key_col, value_col):
    """
    从数据库读取重命名的表，并转化成字典
    :param conn: 数据库连接
    :param db_table_name: 数据库的表名
    :param table_name: 要重命名的表名
    :param key_col: 字典键
    :param value_col: 字典值
    """
    dict_rename = pd.read_sql(
        f"select eng_name, chn_name from {db_table_name} where table_name = '{table_name}'", conn).set_index(key_col).to_dict()[value_col]
    return dict_rename